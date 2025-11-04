"""Encrypted SQLite storage for per-query result history."""

from __future__ import annotations

import csv
import hashlib
import json
import os
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from cryptography.fernet import Fernet
from openpyxl import load_workbook

from opendental_query.constants import DEFAULT_CONFIG_DIR
from opendental_query.renderers.excel_exporter import ExcelExporter
from opendental_query.utils.persist_db import PersistDatabase, _EncryptedDatabaseContext

SUPPORTED_EXCEL_FORMATS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
CSV_ENCODING = "utf-8-sig"


class QueryHistoryDatabase:
    """Manage encrypted SQLite database containing query result history."""

    def __init__(self, config_dir: Path | None = None) -> None:
        self._config_dir = Path(config_dir or DEFAULT_CONFIG_DIR)
        self._config_dir.mkdir(parents=True, exist_ok=True)
        self._db_path = self._config_dir / "query_history.db.enc"
        self._key_path = self._config_dir / "query_history.key"
        self._fernet = Fernet(self._load_key())

    def record_query_result(
        self,
        query_text: str,
        columns: Sequence[str],
        rows: Iterable[Mapping[str, Any]],
        source: str,
        metadata: Mapping[str, Any] | None = None,
    ) -> int:
        """Store rows produced by a query run and log metadata about the run."""
        normalized_query = query_text.strip()
        if not normalized_query:
            raise ValueError("Query text must not be empty")

        rows_list = list(rows)
        if not rows_list:
            return 0

        sanitized_columns = self._sanitize_columns(columns)
        run_timestamp = datetime.now(timezone.utc).isoformat()

        saved_query_name: str | None = None
        if metadata is not None:
            saved_name = metadata.get("saved_query")
            if isinstance(saved_name, str):
                cleaned_name = saved_name.strip()
                if cleaned_name:
                    saved_query_name = cleaned_name

        with _EncryptedDatabaseContext(self._db_path, self._fernet) as conn:
            cursor = conn.cursor()
            self._ensure_schema(cursor)

            query_id = self._hash_query(normalized_query)
            query_record = self._fetch_query_record(cursor, query_id)

            if query_record is None:
                table_name = self._create_query_table(
                    cursor,
                    query_id,
                    normalized_query,
                    columns,
                    sanitized_columns,
                    run_timestamp,
                    preferred_table_name=saved_query_name,
                )
            else:
                table_name = query_record["sanitized_table"]
                stored_columns = json.loads(query_record["columns"])
                if list(columns) != stored_columns:
                    raise ValueError(
                        "Column mismatch detected for existing query history. "
                        "Save this run under a new query instead."
                    )
                sanitized_columns = json.loads(query_record["sanitized_columns"])

            placeholders = ", ".join(["?"] * len(sanitized_columns))
            quoted_columns = ", ".join([f'"{col}"' for col in sanitized_columns])
            insert_sql = f'INSERT INTO "{table_name}" ({quoted_columns}) VALUES ({placeholders})'

            payload = [
                [self._coerce_value(row.get(col)) for col in columns] for row in rows_list
            ]
            cursor.executemany(insert_sql, payload)

            run_metadata = {
                "row_count": len(payload),
                "source": source,
                "columns": list(columns),
            }
            if metadata:
                run_metadata.update(metadata)

            cursor.execute(
                "INSERT INTO query_runs (query_id, run_at, row_count, source, metadata) "
                "VALUES (?, ?, ?, ?, ?)",
                (
                    query_id,
                    run_timestamp,
                    len(payload),
                    source,
                    json.dumps(run_metadata),
                ),
            )

            conn.commit()

        return len(payload)

    def list_queries(self) -> list[dict[str, Any]]:
        """Return metadata for all stored queries."""
        with _EncryptedDatabaseContext(self._db_path, self._fernet) as conn:
            cursor = conn.cursor()
            self._ensure_schema(cursor)
            cursor.execute(
                "SELECT query_id, query_text, sanitized_table, created_at FROM queries ORDER BY created_at"
            )
            columns = [description[0] for description in cursor.description]
            return [dict(zip(columns, row, strict=True)) for row in cursor.fetchall()]

    def saved_query_aliases(self) -> dict[str, str]:
        """Return mapping of query IDs to saved query names derived from run metadata."""
        aliases: dict[str, str] = {}
        with _EncryptedDatabaseContext(self._db_path, self._fernet) as conn:
            cursor = conn.cursor()
            self._ensure_schema(cursor)
            cursor.execute(
                "SELECT query_id, metadata FROM query_runs WHERE metadata IS NOT NULL ORDER BY run_at"
            )
            for query_id, metadata_raw in cursor.fetchall():
                if not metadata_raw:
                    continue
                try:
                    metadata = json.loads(metadata_raw)
                except json.JSONDecodeError:
                    continue
                saved_name = metadata.get("saved_query")
                if isinstance(saved_name, str) and saved_name.strip():
                    aliases[query_id] = saved_name.strip()
        return aliases

    def list_runs(self, query_text: str | None = None) -> list[dict[str, Any]]:
        """Return run metadata for all queries or a specific query."""
        with _EncryptedDatabaseContext(self._db_path, self._fernet) as conn:
            cursor = conn.cursor()
            self._ensure_schema(cursor)

            if query_text is None:
                cursor.execute(
                    "SELECT qr.run_id, qr.query_id, q.query_text, qr.run_at, qr.row_count, qr.source, qr.metadata "
                    "FROM query_runs qr "
                    "JOIN queries q ON q.query_id = qr.query_id "
                    "ORDER BY qr.run_at"
                )
            else:
                query_id = self._hash_query(query_text.strip())
                cursor.execute(
                    "SELECT qr.run_id, qr.query_id, q.query_text, qr.run_at, qr.row_count, qr.source, qr.metadata "
                    "FROM query_runs qr "
                    "JOIN queries q ON q.query_id = qr.query_id "
                    "WHERE qr.query_id = ? "
                    "ORDER BY qr.run_at",
                    (query_id,),
                )

            columns = [description[0] for description in cursor.description]
            return [dict(zip(columns, row, strict=True)) for row in cursor.fetchall()]

    def import_csv(
        self,
        query_text: str,
        csv_path: Path,
        *,
        encoding: str = CSV_ENCODING,
    ) -> int:
        """Load data from a CSV file and record it for the given query."""
        with open(csv_path, newline="", encoding=encoding) as handle:
            reader = csv.DictReader(handle)
            fieldnames = reader.fieldnames
            if not fieldnames:
                raise ValueError("CSV file must include a header row")
            rows = [dict(row) for row in reader]
        if not rows:
            return 0
        metadata = {"source_file": str(csv_path)}
        return self.record_query_result(
            query_text=query_text,
            columns=fieldnames,
            rows=rows,
            source="csv",
            metadata=metadata,
        )

    def import_excel(
        self,
        query_text: str,
        workbook_path: Path,
        *,
        sheet_name: str | None = None,
    ) -> int:
        """Load data from an Excel worksheet and record it for the given query."""
        path = Path(workbook_path)
        if path.suffix.lower() not in SUPPORTED_EXCEL_FORMATS:
            raise ValueError(f"Unsupported Excel format: {path.suffix}")

        wb = load_workbook(filename=path, read_only=True, data_only=True)
        try:
            try:
                ws = wb[sheet_name] if sheet_name else wb.active
            except KeyError as exc:
                raise ValueError(f"Worksheet '{sheet_name}' not found") from exc
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration as exc:
                raise ValueError("Worksheet does not contain any rows") from exc
            if header_row is None:
                raise ValueError("Worksheet header row is empty")

            columns = [self._normalize_header(cell, index) for index, cell in enumerate(header_row)]
            if not any(columns):
                raise ValueError("Worksheet header row must include at least one column name")

            records = []
            for row in rows_iter:
                if row is None:
                    continue
                if all(value is None for value in row):
                    continue
                record = {
                    columns[idx]: row[idx] if idx < len(row) else None for idx in range(len(columns))
                }
                records.append(record)
        finally:
            wb.close()

        if not records:
            return 0

        metadata = {"source_file": str(path)}
        if sheet_name:
            metadata["sheet_name"] = sheet_name

        return self.record_query_result(
            query_text=query_text,
            columns=columns,
            rows=records,
            source="excel",
            metadata=metadata,
        )

    def export_query_to_csv(self, query_text: str, output_path: Path) -> int:
        """Export stored query rows to a CSV file."""
        columns, rows = self._load_query_rows(query_text)

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with output_path.open("w", newline="", encoding=CSV_ENCODING) as handle:
            writer = csv.writer(handle)
            writer.writerow(columns)
            for row in rows:
                writer.writerow(
                    [
                        "" if (idx >= len(row) or row[idx] is None) else str(row[idx])
                        for idx in range(len(columns))
                    ]
                )

        return len(rows)

    def export_query_to_excel(
        self,
        query_text: str,
        output_path: Path | None = None,
    ) -> tuple[Path | None, int]:
        """Export stored query rows to an Excel workbook."""
        columns, rows = self._load_query_rows(query_text)
        data = [
            {columns[idx]: (row[idx] if idx < len(row) and row[idx] is not None else "") for idx in range(len(columns))}
            for row in rows
        ]

        if not data:
            return None, 0

        exporter = ExcelExporter()

        if output_path is None:
            exported_path = exporter.export(data, preserve_order=True)
            return exported_path, len(data)

        output_path = Path(output_path)
        if output_path.suffix:
            if output_path.suffix.lower() != ".xlsx":
                raise ValueError("Output path must be a directory or a .xlsx file.")
            output_path.parent.mkdir(parents=True, exist_ok=True)
            temp_path = exporter.export(data, output_dir=output_path.parent, preserve_order=True)
            exported_path = temp_path.replace(output_path)
            return exported_path, len(data)

        output_path.mkdir(parents=True, exist_ok=True)
        exported_path = exporter.export(data, output_dir=output_path, preserve_order=True)
        return exported_path, len(data)

    def delete_query_history(self, query_text: str) -> bool:
        """Delete persisted history for the specified query text."""
        normalized_query = query_text.strip()
        if not normalized_query:
            raise ValueError("Query text must not be empty")

        with _EncryptedDatabaseContext(self._db_path, self._fernet) as conn:
            cursor = conn.cursor()
            self._ensure_schema(cursor)

            query_id = self._hash_query(normalized_query)
            query_record = self._fetch_query_record(cursor, query_id)
            if query_record is None:
                return False

            sanitized_table = query_record["sanitized_table"]
            cursor.execute(f'DROP TABLE IF EXISTS "{sanitized_table}"')
            cursor.execute("DELETE FROM query_runs WHERE query_id = ?", (query_id,))
            cursor.execute("DELETE FROM queries WHERE query_id = ?", (query_id,))
            conn.commit()
        return True

    def _create_query_table(
        self,
        cursor: sqlite3.Cursor,
        query_id: str,
        query_text: str,
        columns: Sequence[str],
        sanitized_columns: Sequence[str],
        created_at: str,
        *,
        preferred_table_name: str | None = None,
    ) -> str:
        table_name = self._generate_table_identifier(
            cursor,
            query_id,
            preferred_table_name,
        )
        column_defs = ", ".join([f'"{col}" TEXT' for col in sanitized_columns])
        cursor.execute(f'CREATE TABLE "{table_name}" ({column_defs})')
        cursor.execute(
            "INSERT INTO queries (query_id, query_text, sanitized_table, columns, sanitized_columns, created_at) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (
                query_id,
                query_text,
                table_name,
                json.dumps(list(columns)),
                json.dumps(list(sanitized_columns)),
                created_at,
            ),
        )
        return table_name

    def _generate_table_identifier(
        self,
        cursor: sqlite3.Cursor,
        query_id: str,
        preferred_table_name: str | None,
    ) -> str:
        if preferred_table_name:
            sanitized = self._sanitize_identifier(preferred_table_name)
            if sanitized:
                return self._ensure_unique_table_name(cursor, sanitized)

        default_candidate = f"q_{query_id[:16]}"
        return self._ensure_unique_table_name(cursor, default_candidate)

    @classmethod
    def _ensure_unique_table_name(cls, cursor: sqlite3.Cursor, base_name: str) -> str:
        candidate = base_name
        suffix = 1
        while cls._table_exists(cursor, candidate):
            candidate = f"{base_name}_{suffix}"
            suffix += 1
        return candidate

    @staticmethod
    def _table_exists(cursor: sqlite3.Cursor, table_name: str) -> bool:
        cursor.execute(
            "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
            (table_name,),
        )
        return cursor.fetchone() is not None

    @staticmethod
    def _sanitize_identifier(name: str) -> str:
        return PersistDatabase._sanitize_identifier(name)

    def _ensure_schema(self, cursor: sqlite3.Cursor) -> None:
        cursor.execute(
            "CREATE TABLE IF NOT EXISTS queries ("
            "query_id TEXT PRIMARY KEY,"
            "query_text TEXT NOT NULL,"
            "sanitized_table TEXT NOT NULL,"
            "columns TEXT NOT NULL,"
            "sanitized_columns TEXT NOT NULL,"
            "created_at TEXT NOT NULL"
            ")"
        )
        cursor.execute(
            "CREATE TABLE IF NOT EXISTS query_runs ("
            "run_id INTEGER PRIMARY KEY AUTOINCREMENT,"
            "query_id TEXT NOT NULL,"
            "run_at TEXT NOT NULL,"
            "row_count INTEGER NOT NULL,"
            "source TEXT NOT NULL,"
            "metadata TEXT,"
            "FOREIGN KEY(query_id) REFERENCES queries(query_id)"
            ")"
        )

    def _fetch_query_record(self, cursor: sqlite3.Cursor, query_id: str) -> dict[str, Any] | None:
        cursor.execute(
            "SELECT query_id, query_text, sanitized_table, columns, sanitized_columns FROM queries WHERE query_id = ?",
            (query_id,),
        )
        row = cursor.fetchone()
        if row is None:
            return None
        keys = [description[0] for description in cursor.description]
        return dict(zip(keys, row, strict=True))

    @staticmethod
    def _coerce_value(value: Any) -> str:
        if value is None:
            return ""
        return str(value)

    @staticmethod
    def _hash_query(query_text: str) -> str:
        return hashlib.sha256(query_text.encode("utf-8")).hexdigest()

    @staticmethod
    def _sanitize_columns(columns: Sequence[str]) -> list[str]:
        sanitized: list[str] = []
        seen: set[str] = set()
        for column in columns:
            base = QueryHistoryDatabase._sanitize_identifier(column)
            candidate = base
            index = 1
            while candidate in seen:
                candidate = f"{base}_{index}"
                index += 1
            sanitized.append(candidate)
            seen.add(candidate)
        return sanitized

    def _load_key(self) -> bytes:
        if self._key_path.exists():
            key = self._key_path.read_bytes()
            if key:
                return key

        key = Fernet.generate_key()
        self._key_path.write_bytes(key)
        try:
            if os.name != "nt":
                os.chmod(self._key_path, 0o600)
        except OSError:
            pass
        return key

    @staticmethod
    def _normalize_header(value: Any, index: int) -> str:
        if value is None:
            return f"column_{index+1}"
        text = str(value).strip()
        return text or f"column_{index+1}"

    def _load_query_rows(self, query_text: str) -> tuple[list[str], list[tuple[Any, ...]]]:
        normalized_query = query_text.strip()
        if not normalized_query:
            raise ValueError("Query text must not be empty")

        with _EncryptedDatabaseContext(self._db_path, self._fernet) as conn:
            cursor = conn.cursor()
            self._ensure_schema(cursor)

            query_id = self._hash_query(normalized_query)
            query_record = self._fetch_query_record(cursor, query_id)
            if query_record is None:
                raise ValueError("No stored history found for the provided query.")

            table_name = query_record["sanitized_table"]
            columns = json.loads(query_record["columns"])
            sanitized_columns = json.loads(query_record["sanitized_columns"])

            quoted_columns = ", ".join([f'"{col}"' for col in sanitized_columns])
            select_sql = f'SELECT rowid, {quoted_columns} FROM "{table_name}" ORDER BY rowid'
            cursor.execute(select_sql)
            rows_with_id = cursor.fetchall()

            cursor.execute(
                "SELECT run_at, row_count FROM query_runs WHERE query_id = ? ORDER BY run_at",
                (query_id,),
            )
            run_batches = cursor.fetchall()

        office_index = None
        try:
            office_index = columns.index("Office")
        except ValueError:
            office_index = None

        ordered_rows: list[tuple[Any, ...]] = []
        current_index = 0
        total_rows = [tuple(row[1:]) for row in rows_with_id]

        for run_at, row_count in run_batches:
            batch: list[tuple[Any, ...]] = []
            for _ in range(int(row_count)):
                if current_index >= len(total_rows):
                    break
                batch.append(total_rows[current_index])
                current_index += 1
            if office_index is not None:
                batch.sort(
                    key=lambda row: (
                        row[office_index] if row[office_index] is not None else ""
                    )
                )
            ordered_rows.extend(batch)

        if current_index < len(total_rows):
            remaining = total_rows[current_index:]
            if office_index is not None:
                remaining.sort(
                    key=lambda row: row[office_index] if row[office_index] is not None else ""
                )
            ordered_rows.extend(remaining)

        return columns, ordered_rows
