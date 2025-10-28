"""
Excel exporter for query results.

Creates secure-by-default XLSX workbooks with:
- Random filename generation (opendental_query_{token}_{timestamp}.xlsx)
- Table formatting (White, Table Style Light 8)
- Auto-sized columns with configurable width bounds
- Downloads folder detection with fallback to cwd
- Optional post-export encryption hook
"""

from __future__ import annotations

import os
import secrets
import shlex
import subprocess
from datetime import date, datetime, time
from decimal import Decimal
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from opendental_query.constants import DEFAULT_CONFIG_DIR
from opendental_query.utils.formatting import format_cell_value


class ExcelExporter:
    """
    Export query results to XLSX files with secure defaults and table styling.
    """

    TABLE_STYLE_NAME = "TableStyleLight8"
    TABLE_DISPLAY_PREFIX = "Results_"
    MIN_COLUMN_WIDTH = 12
    MAX_COLUMN_WIDTH = 60

    def export(
        self,
        rows: list[dict[str, Any]],
        output_dir: Path | None = None,
        *,
        preserve_order: bool = False,
    ) -> Path:
        if not rows:
            raise ValueError("Cannot export empty rows list - no data to write")

        target_dir = self._resolve_output_dir(output_dir)
        self._ensure_secure_directory(target_dir)
        target_dir.mkdir(parents=True, exist_ok=True)
        self._set_permissions(target_dir, dir_mode=0o700)

        filename = self._generate_filename()
        filepath = (target_dir / filename).resolve()

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Results"

        fieldnames = list(rows[0].keys())
        sorted_rows = list(rows) if preserve_order else sorted(rows, key=lambda row: str(row.get("Office", "")))
        worksheet.append(fieldnames)

        for row_index, row in enumerate(sorted_rows, start=2):
            cleaned_row = {key: format_cell_value(value) for key, value in row.items()}
            for col_index, key in enumerate(fieldnames, start=1):
                raw_value = row.get(key)
                excel_value, number_format = self._prepare_excel_value(raw_value)
                display_value = cleaned_row.get(key, "")
                if excel_value is None:
                    cell_value = self._sanitize_excel_text(display_value)
                else:
                    if isinstance(excel_value, str):
                        cell_value = self._sanitize_excel_text(excel_value)
                    else:
                        cell_value = excel_value
                cell = worksheet.cell(row=row_index, column=col_index, value=cell_value)
                if number_format is not None and excel_value is not None:
                    cell.number_format = number_format

        self._apply_table_style(worksheet, fieldnames, row_count=len(rows))
        self._auto_fit_columns(worksheet, fieldnames)
        self._disable_wrap_text(worksheet)

        workbook.save(filepath)

        self._set_permissions(filepath, file_mode=0o600)
        return self._maybe_encrypt_file(filepath)

    def _resolve_output_dir(self, output_dir: Path | None) -> Path:
        return Path(output_dir) if output_dir is not None else self._get_downloads_folder()

    def _ensure_secure_directory(self, output_dir: Path) -> None:
        if os.getenv("SPEC_KIT_ALLOW_UNSAFE_EXPORTS") == "1":
            return

        allowed_roots = [self._get_downloads_folder().resolve(), DEFAULT_CONFIG_DIR.resolve()]
        env_root = os.getenv("SPEC_KIT_EXPORT_ROOT")
        if env_root:
            allowed_roots.append(Path(env_root).expanduser().resolve())

        resolved_dir = output_dir.resolve()
        for root in allowed_roots:
            try:
                if resolved_dir == root or resolved_dir.is_relative_to(root):
                    return
            except AttributeError:
                try:
                    resolved_dir.relative_to(root)
                    return
                except ValueError:
                    continue

        raise ValueError(
            f"Export directory {resolved_dir} is outside allowed roots. "
            "Set SPEC_KIT_EXPORT_ROOT to an approved location or set "
            "SPEC_KIT_ALLOW_UNSAFE_EXPORTS=1 to override (not recommended)."
        )

    def _set_permissions(
        self,
        path: Path,
        *,
        dir_mode: int | None = None,
        file_mode: int | None = None,
    ) -> None:
        if os.name == "nt":
            return
        mode = dir_mode if path.is_dir() else file_mode
        if mode is None:
            return
        try:
            os.chmod(path, mode)
        except PermissionError:
            pass

    def _generate_filename(self) -> str:
        token = secrets.token_hex(4)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"opendental_query_{token}_{timestamp}.xlsx"

    def _generate_table_name(self) -> str:
        token = secrets.token_hex(3)
        return f"{self.TABLE_DISPLAY_PREFIX}{token}"

    def _apply_table_style(self, worksheet, fieldnames: list[str], row_count: int) -> None:
        if not fieldnames or row_count < 0:
            return

        last_column_letter = get_column_letter(len(fieldnames))
        last_row = row_count + 1  # include header
        table_ref = f"A1:{last_column_letter}{last_row}"

        table = Table(displayName=self._generate_table_name(), ref=table_ref)
        table_style = TableStyleInfo(
            name=self.TABLE_STYLE_NAME,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = table_style
        worksheet.add_table(table)

    def _auto_fit_columns(self, worksheet, fieldnames: list[str]) -> None:
        for idx, _ in enumerate(fieldnames, start=1):
            column_letter = get_column_letter(idx)
            max_length = 0
            for cell in worksheet[column_letter]:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            adjusted_width = min(
                max(self.MIN_COLUMN_WIDTH, max_length + 2),
                self.MAX_COLUMN_WIDTH,
            )
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _disable_wrap_text(self, worksheet) -> None:
        no_wrap = Alignment(wrap_text=False)
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = no_wrap

    def _sanitize_excel_text(self, value: str) -> str:
        if not isinstance(value, str):
            return value

        if not value:
            return value

        first = value[0]
        if first in ("=", "+", "-", "@", "\t", "\r", "\n"):
            return f"'{value}"
        return value

    _CURRENCY_PATTERN = re.compile(
        r"""
        ^\s*
        (?P<open_paren>\()?          # optional opening parenthesis for negatives
        (?P<sign>-)?                 # optional leading minus
        \$\s*
        (?P<number>[0-9.,]+)         # digits, commas, decimals
        \)?\s*$
        """,
        re.VERBOSE,
    )

    _PERCENT_PATTERN = re.compile(r"^\s*(?P<value>-?[0-9.,]+)\s*%\s*$")
    _GROUPED_NUMBER_PATTERN = re.compile(r"^\s*-?[0-9.,]+(?:\.[0-9]+)?\s*$")

    def _prepare_excel_value(self, raw_value: Any) -> tuple[Any | None, str | None]:
        """Coerce values and provide number formats for Excel."""
        dt = self._try_parse_datetime(raw_value)
        if dt is not None:
            if self._is_midnight(dt):
                return dt, "mm/dd/yyyy"
            return dt, "mm/dd/yyyy hh:mm:ss"

        if not isinstance(raw_value, str):
            return raw_value, None

        stripped = raw_value.strip()
        if stripped == "":
            return "", None

        numeric_portion = stripped.lstrip("+-")
        if numeric_portion.startswith("0") and len(numeric_portion) > 1 and numeric_portion.isdigit():
            return None, None

        currency_match = self._CURRENCY_PATTERN.match(stripped)
        if currency_match:
            numeric = currency_match.group("number").replace(",", "")
            decimals = self._count_decimals(numeric)
            amount = Decimal(numeric)
            if currency_match.group("open_paren") or currency_match.group("sign"):
                amount = -amount
            number_format = self._build_number_format("$#,##0", decimals)
            return amount, number_format

        percent_match = self._PERCENT_PATTERN.match(stripped)
        if percent_match:
            numeric = percent_match.group("value").replace(",", "")
            decimals = self._count_decimals(numeric)
            try:
                amount = Decimal(numeric) / Decimal(100)
            except Exception:
                return None, None
            number_format = self._build_number_format("0", decimals) + "%"
            return amount, number_format

        if self._GROUPED_NUMBER_PATTERN.match(stripped):
            cleaned = stripped.replace(",", "")
            decimals = self._count_decimals(cleaned)
            try:
                if decimals == 0:
                    amount = int(cleaned)
                else:
                    amount = float(cleaned)
            except ValueError:
                return None, None
            number_format = None
            if "," in stripped:
                number_format = self._build_number_format("#,##0", decimals)
            return amount, number_format

        if stripped.isdigit() or (stripped.startswith("-") and stripped[1:].isdigit()):
            try:
                return int(stripped), None
            except ValueError:
                return None, None

        try:
            float_value = float(stripped)
        except ValueError:
            return None, None

        if stripped.lower().startswith(("nan", "inf")):
            return None, None
        return float_value, None


    def _try_parse_datetime(self, value: Any) -> datetime | None:
        if isinstance(value, datetime):
            return value.replace(tzinfo=None)
        if isinstance(value, date):
            return datetime.combine(value, time.min)
        if not isinstance(value, str):
            return None
        stripped = value.strip()
        if not stripped:
            return None
        try:
            return datetime.fromisoformat(stripped)
        except ValueError:
            pass
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
            "%m-%d-%Y",
            "%m/%d/%Y",
            "%m/%d/%Y %H:%M:%S",
            "%m-%d-%Y %H:%M:%S",
        ):
            try:
                return datetime.strptime(stripped, fmt)
            except ValueError:
                continue
        return None

    @staticmethod
    def _is_midnight(dt: datetime) -> bool:
        return dt.time().replace(microsecond=0) == time(0, 0, 0)

    @staticmethod
    def _count_decimals(number_text: str) -> int:
        if "." not in number_text:
            return 0
        return len(number_text.split(".", 1)[1])

    @staticmethod
    def _build_number_format(prefix: str, decimals: int) -> str:
        if decimals <= 0:
            return prefix
        return f"{prefix}." + ("0" * decimals)

    def _get_downloads_folder(self) -> Path:
        downloads = Path.home() / "Downloads"
        return downloads if downloads.exists() and downloads.is_dir() else Path.cwd()

    def _maybe_encrypt_file(self, filepath: Path) -> Path:
        command_template = os.getenv("SPEC_KIT_EXPORT_ENCRYPTION_COMMAND")
        if not command_template:
            return filepath

        command = command_template.format(input=str(filepath))
        try:
            args = shlex.split(command, posix=(os.name != "nt"))
            subprocess.run(args, check=True)
        except Exception as exc:
            raise RuntimeError(f"Export encryption command failed: {exc}") from exc
        return filepath

