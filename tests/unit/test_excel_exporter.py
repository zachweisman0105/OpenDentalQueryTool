"""
Unit tests for ExcelExporter.

Validates workbook creation, table styling, column sizing, and security controls.
"""

from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path
from datetime import datetime
from typing import Any
from unittest.mock import patch

import openpyxl
import pytest

from opendental_query.renderers.excel_exporter import ExcelExporter


@pytest.fixture
def allow_unrestricted_exports(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("SPEC_KIT_ALLOW_UNSAFE_EXPORTS", "1")
    yield
    monkeypatch.delenv("SPEC_KIT_ALLOW_UNSAFE_EXPORTS", raising=False)
    monkeypatch.delenv("SPEC_KIT_EXPORT_ENCRYPTION_COMMAND", raising=False)
    monkeypatch.delenv("SPEC_KIT_EXPORT_ROOT", raising=False)


def _load_workbook(path: Path) -> openpyxl.Workbook:
    return openpyxl.load_workbook(path)


class TestFilenameGeneration:
    """Test random filename generation."""

    def test_generates_unique_filenames(self, allow_unrestricted_exports: None) -> None:
        exporter = ExcelExporter()
        filenames = [exporter._generate_filename() for _ in range(10)]
        assert len(set(filenames)) == 10
        assert all(name.endswith(".xlsx") for name in filenames)

    def test_filename_format_includes_timestamp(self, allow_unrestricted_exports: None) -> None:
        exporter = ExcelExporter()
        filename = exporter._generate_filename()
        assert filename.startswith("opendental_query_")
        assert filename.endswith(".xlsx")
        parts = filename.replace(".xlsx", "").split("_")
        assert len(parts) >= 3


class TestDownloadsFolderDetection:
    """Test Downloads folder detection and fallback."""

    def test_detects_downloads_folder(self, allow_unrestricted_exports: None) -> None:
        exporter = ExcelExporter()
        downloads_path = exporter._get_downloads_folder()
        assert isinstance(downloads_path, Path)
        assert "Downloads" in str(downloads_path) or downloads_path == Path.cwd()

    def test_falls_back_to_cwd_if_downloads_not_found(self, allow_unrestricted_exports: None) -> None:
        exporter = ExcelExporter()
        with patch("pathlib.Path.home") as mock_home:
            fake_home = Path(tempfile.gettempdir()) / "fake_home"
            mock_home.return_value = fake_home
            downloads_path = exporter._get_downloads_folder()
            assert downloads_path == Path.cwd()


class TestExcelWorkbook:
    """Tests relating to the generated workbook."""

    def test_export_writes_rows(self, allow_unrestricted_exports: None) -> None:
        exporter = ExcelExporter()
        rows = [
            {"Office": "office1", "Name": "Smith"},
            {"Office": "office2", "Name": "Jones"},
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir)
            filepath = exporter.export(rows, output_dir=output_dir)
            assert filepath.suffix == ".xlsx"
            assert filepath.exists()

            workbook = _load_workbook(filepath)
            sheet = workbook.active
            values = list(sheet.iter_rows(values_only=True))
            assert values[0] == ("Office", "Name")
            assert values[1] == ("office1", "Smith")
            assert values[2] == ("office2", "Jones")

    def test_table_style_applied(self, allow_unrestricted_exports: None) -> None:
        exporter = ExcelExporter()
        rows = [{"Office": "office1", "Name": "Smith"}]
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = exporter.export(rows, output_dir=Path(tmpdir))
            workbook = _load_workbook(filepath)
            sheet = workbook.active
            assert sheet.tables, "Expected a formatted table to be present"
            table = next(iter(sheet.tables.values()))
            assert table.tableStyleInfo.name == exporter.TABLE_STYLE_NAME

    def test_column_widths_bounded(self, allow_unrestricted_exports: None) -> None:
        exporter = ExcelExporter()
        rows = [
            {"Short": "abc", "Long": "X" * 256},
            {"Short": "short text", "Long": "medium"},
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = exporter.export(rows, output_dir=Path(tmpdir))
            workbook = _load_workbook(filepath)
            sheet = workbook.active

            width_short = sheet.column_dimensions["A"].width
            width_long = sheet.column_dimensions["B"].width

            assert width_short is not None
            assert width_long is not None
            assert exporter.MIN_COLUMN_WIDTH <= width_short <= exporter.MAX_COLUMN_WIDTH
            assert width_long == pytest.approx(exporter.MAX_COLUMN_WIDTH)

    def test_wrap_text_disabled(self, allow_unrestricted_exports: None) -> None:
        exporter = ExcelExporter()
        rows = [{"Col": "Wrapped text should be disabled"}]
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = exporter.export(rows, output_dir=Path(tmpdir))
            workbook = _load_workbook(filepath)
            sheet = workbook.active
            for row in sheet.iter_rows():
                for cell in row:
                    assert cell.alignment is not None
                    assert cell.alignment.wrapText in (None, False)

    def test_formats_midnight_timestamp_without_time_component(
        self, allow_unrestricted_exports: None
    ) -> None:
        exporter = ExcelExporter()
        rows = [
            {"ApptDate": "2025-10-25T00:00:00"},
            {"ApptDate": "2025-10-25T15:30:45"},
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = exporter.export(rows, output_dir=Path(tmpdir))
            workbook = _load_workbook(filepath)
            sheet = workbook.active
            cell_midnight = sheet.cell(row=2, column=1)
            cell_time = sheet.cell(row=3, column=1)
            assert cell_midnight.value == datetime(2025, 10, 25, 0, 0)
            assert cell_midnight.number_format == "mm/dd/yyyy"
            assert cell_time.value == datetime(2025, 10, 25, 15, 30, 45)
            assert cell_time.number_format == "mm/dd/yyyy hh:mm:ss"

    def test_handles_none_and_missing_values(self, allow_unrestricted_exports: None) -> None:
        exporter = ExcelExporter()
        rows = [
            {"Office": "office1", "PatNum": None, "LName": "Smith"},
            {"Office": "office2", "LName": "Jones"},
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = exporter.export(rows, output_dir=Path(tmpdir))
            workbook = _load_workbook(filepath)
            sheet = workbook.active
            values = list(sheet.iter_rows(values_only=True))
            assert values[1][0] == "office1"
            assert values[1][1] in ("", None)
            assert values[1][2] == "Smith"
            assert values[2][0] == "office2"
            assert values[2][1] in ("", None)
            assert values[2][2] == "Jones"

    def test_currency_strings_format_as_currency(self, allow_unrestricted_exports: None) -> None:
        exporter = ExcelExporter()
        rows = [
            {"Amount": "$1,234.50"},
            {"Amount": "($2,000.00)"},
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = exporter.export(rows, output_dir=Path(tmpdir))
            workbook = _load_workbook(filepath)
            sheet = workbook.active
            cell_row1 = sheet.cell(row=2, column=1)
            cell_row2 = sheet.cell(row=3, column=1)
            assert float(cell_row1.value) == pytest.approx(1234.50)
            assert float(cell_row2.value) == pytest.approx(-2000.00)
            assert cell_row1.number_format == "$#,##0.00"
            assert cell_row2.number_format == "$#,##0.00"

    def test_percentage_strings_format_as_percentage(self, allow_unrestricted_exports: None) -> None:
        exporter = ExcelExporter()
        rows = [
            {"Rate": "12.5%"},
            {"Rate": "-3%"},
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = exporter.export(rows, output_dir=Path(tmpdir))
            workbook = _load_workbook(filepath)
            sheet = workbook.active
            cell_row1 = sheet.cell(row=2, column=1)
            cell_row2 = sheet.cell(row=3, column=1)
            assert float(cell_row1.value) == pytest.approx(0.125)
            assert float(cell_row2.value) == pytest.approx(-0.03)
            assert cell_row1.number_format == "0.0%"
            assert cell_row2.number_format == "0%"

    def test_numeric_strings_coerce_to_numbers(self, allow_unrestricted_exports: None) -> None:
        exporter = ExcelExporter()
        rows = [
            {"Count": "5", "Price": "12.50", "Code": "0012"},
            {"Count": "-3", "Price": "7.0", "Code": "A123"},
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = exporter.export(rows, output_dir=Path(tmpdir))
            workbook = _load_workbook(filepath)
            sheet = workbook.active
            cells_row1 = list(sheet.iter_rows(min_row=2, max_row=2))[0]
            assert cells_row1[0].value == 5
            assert isinstance(cells_row1[0].value, int)
            assert cells_row1[1].value == pytest.approx(12.50)
            # Leading zeros preserved as text
            assert cells_row1[2].value == "0012"

            cells_row2 = list(sheet.iter_rows(min_row=3, max_row=3))[0]
            assert cells_row2[0].value == -3
            assert isinstance(cells_row2[0].value, int)
            assert cells_row2[1].value == pytest.approx(7.0)
            assert cells_row2[2].value == "A123"

    def test_sanitizes_formula_like_strings(self, allow_unrestricted_exports: None) -> None:
        exporter = ExcelExporter()
        rows = [
            {"Col": "=SUM(1,2)", "Notes": "+Important"},
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = exporter.export(rows, output_dir=Path(tmpdir))
            workbook = _load_workbook(filepath)
            sheet = workbook.active
            cells = list(sheet.iter_rows(min_row=2, values_only=False))[0]
            for cell in cells:
                assert cell.data_type == "s"
                assert not (cell.value or "").startswith("=")
                assert not (cell.value or "").startswith("+")


class TestExportMethod:
    """Test export method behavior."""

    def test_returns_filepath(self, allow_unrestricted_exports: None) -> None:
        exporter = ExcelExporter()
        rows = [{"Office": "office1", "Name": "Smith"}]
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir)
            filepath = exporter.export(rows, output_dir=output_dir)
            assert isinstance(filepath, Path)
            assert filepath.exists()
            assert filepath.parent == output_dir

    def test_uses_downloads_folder_by_default(self, allow_unrestricted_exports: None) -> None:
        exporter = ExcelExporter()
        rows = [{"Office": "office1", "Name": "Smith"}]
        with patch.object(exporter, "_get_downloads_folder") as mock_downloads:
            mock_downloads_path = Path(tempfile.gettempdir()) / "Downloads"
            mock_downloads_path.mkdir(exist_ok=True)
            mock_downloads.return_value = mock_downloads_path
            filepath = exporter.export(rows)
            assert filepath.parent == mock_downloads_path
            if filepath.exists():
                filepath.unlink()

    def test_creates_output_directory_if_missing(self, allow_unrestricted_exports: None) -> None:
        exporter = ExcelExporter()
        rows = [{"Office": "office1", "Name": "Smith"}]
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir) / "subdir" / "nested"
            assert not output_dir.exists()
            filepath = exporter.export(rows, output_dir=output_dir)
            assert output_dir.exists()
            assert filepath.exists()


class TestEdgeCases:
    """Test edge cases and error handling."""

    def test_handles_empty_rows_list(self, allow_unrestricted_exports: None) -> None:
        exporter = ExcelExporter()
        rows: list[dict[str, Any]] = []
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir)
            with pytest.raises(ValueError, match="empty|no data"):
                exporter.export(rows, output_dir=output_dir)


class TestExportSecurity:
    """Tests enforcing secure export destinations and optional encryption."""

    def test_rejects_directory_outside_allowed_roots(
        self, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
    ) -> None:
        monkeypatch.delenv("SPEC_KIT_ALLOW_UNSAFE_EXPORTS", raising=False)
        monkeypatch.delenv("SPEC_KIT_EXPORT_ROOT", raising=False)
        exporter = ExcelExporter()
        rows = [{"Office": "office1", "Name": "Smith"}]
        with pytest.raises(ValueError):
            exporter.export(rows, output_dir=tmp_path / "untrusted")

    def test_allows_configured_export_root(
        self, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
    ) -> None:
        safe_root = tmp_path / "approved"
        safe_root.mkdir()
        monkeypatch.setenv("SPEC_KIT_EXPORT_ROOT", str(safe_root))
        exporter = ExcelExporter()
        rows = [{"Office": "office1", "Name": "Jones"}]
        filepath = exporter.export(rows, output_dir=safe_root)
        assert filepath.exists()

    def test_encryption_command_runs(
        self, monkeypatch: pytest.MonkeyPatch, tmp_path: Path, allow_unrestricted_exports: None
    ) -> None:
        encrypt_script = tmp_path / "encrypt.py"
        encrypt_script.write_text(
            "import pathlib\n"
            "import sys\n"
            "path = pathlib.Path(sys.argv[1])\n"
            "flag = path.with_suffix(path.suffix + '.flag')\n"
            "flag.write_text('encrypted')\n"
        )
        monkeypatch.setenv(
            "SPEC_KIT_EXPORT_ENCRYPTION_COMMAND",
            f"{sys.executable} {encrypt_script} {{input}}",
        )
        exporter = ExcelExporter()
        rows = [{"Office": "office1", "Name": "Doe"}]
        filepath = exporter.export(rows, output_dir=tmp_path)
        flag_file = filepath.with_suffix(filepath.suffix + ".flag")
        assert flag_file.exists()
