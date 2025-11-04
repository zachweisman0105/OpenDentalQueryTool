"""Tests for history CLI commands and shortcuts."""

from __future__ import annotations

import re
import sys
from pathlib import Path
from unittest.mock import patch

import pytest
from click.testing import CliRunner
from openpyxl import Workbook, load_workbook

from opendental_query.cli.history_cmd import history_group
from opendental_query.cli.shortcuts import (
    query_table_shortcut,
    table_export_shortcut,
    table_import_shortcut,
    table_list_shortcut,
    table_delete_shortcut,
    update_table_shortcut,
)
from opendental_query.utils.query_history_db import QueryHistoryDatabase
from opendental_query.utils.saved_queries import SavedQueryLibrary


def _write_excel(path: Path, rows: list[tuple[str, str]], sheet_name: str = "Data") -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(["Office", "Value"])
    for office, value in rows:
        sheet.append([office, value])
    workbook.save(path)
    workbook.close()


def test_create_table_invokes_run_history(tmp_path: Path) -> None:
    library = SavedQueryLibrary(tmp_path)
    library.save_query(name="monthly", sql="SELECT 1")

    runner = CliRunner()
    with patch("opendental_query.cli.history_cmd.run_query_history") as mock_run:
        result = runner.invoke(
            history_group,
            ["create-table"],
            obj={"config_dir": tmp_path},
            input="monthly\n",
        )

    assert result.exit_code == 0
    mock_run.assert_called_once()
    kwargs = mock_run.call_args.kwargs
    assert kwargs["saved_query"] == "monthly"
    assert kwargs["sql"] is None


def test_create_table_requires_saved_query(tmp_path: Path) -> None:
    runner = CliRunner()
    result = runner.invoke(
        history_group,
        ["create-table"],
        obj={"config_dir": tmp_path},
    )
    assert result.exit_code != 0
    assert "No saved queries" in result.output


def test_query_table_shortcut_invokes_cli(monkeypatch) -> None:
    monkeypatch.setattr(sys, "argv", ["QueryTable"])
    with patch("opendental_query.cli.shortcuts.cli") as mock_cli:
        query_table_shortcut()

    mock_cli.assert_called_once_with(obj={})
    assert sys.argv[1:3] == ["history", "create-table"]


def test_export_history_writes_excel(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("SPEC_KIT_ALLOW_UNSAFE_EXPORTS", "1")
    library = SavedQueryLibrary(tmp_path)
    saved = library.save_query(name="monthly", sql="SELECT Office, Value FROM sample")

    db = QueryHistoryDatabase(tmp_path)
    db.record_query_result(
        saved.sql,
        ["Office", "Value"],
        [{"Office": "A1", "Value": "10"}, {"Office": "B2", "Value": "12"}],
        source="test",
        metadata=None,
    )

    runner = CliRunner()
    result = runner.invoke(
        history_group,
        ["export", "--saved-query", "monthly", "--output", str(tmp_path / "exports")],
        obj={"config_dir": tmp_path},
    )

    assert result.exit_code == 0
    match = re.search(r"Exported \d+ row\(s\) .* to (.+)\.", result.output)
    assert match
    exported_path = Path(match.group(1).strip())
    assert exported_path.suffix == ".xlsx"
    assert exported_path.exists()
    workbook = load_workbook(exported_path)
    sheet = workbook.active
    assert sheet.title == "Results"
    values = list(sheet.iter_rows(values_only=True))
    assert values[0][:2] == ("Office", "Value")
    data_rows = values[1:]
    extracted = {(row[0], str(row[1])) for row in data_rows}
    assert extracted == {("A1", "10"), ("B2", "12")}
    workbook.close()


def test_export_history_requires_saved_rows(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setenv("SPEC_KIT_ALLOW_UNSAFE_EXPORTS", "1")
    runner = CliRunner()
    result = runner.invoke(
        history_group,
        ["export", "--saved-query", "empty", "--output", str(tmp_path / "out.xlsx")],
        obj={"config_dir": tmp_path},
    )

    assert result.exit_code != 0
    assert "No persisted query history found" in result.output


def test_table_export_shortcut_invokes_cli(monkeypatch) -> None:
    monkeypatch.setattr(sys, "argv", ["TableExport"])
    with patch("opendental_query.cli.shortcuts.cli") as mock_cli:
        table_export_shortcut()

    mock_cli.assert_called_once_with(obj={})
    assert sys.argv[1:3] == ["history", "export"]


def test_update_table_shortcut_invokes_cli(monkeypatch) -> None:
    monkeypatch.setattr(sys, "argv", ["UpdateTable"])
    with patch("opendental_query.cli.shortcuts.cli") as mock_cli:
        update_table_shortcut()

    mock_cli.assert_called_once_with(obj={})
    assert sys.argv[1:3] == ["history", "run"]


def test_import_table_command_replace(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("SPEC_KIT_ALLOW_UNSAFE_EXPORTS", "1")
    library = SavedQueryLibrary(tmp_path)
    saved = library.save_query(name="monthly", sql="SELECT Office, Value FROM sample")

    db = QueryHistoryDatabase(tmp_path)
    db.record_query_result(
        saved.sql,
        ["Office", "Value"],
        [{"Office": "OLD", "Value": "1"}],
        source="test",
        metadata=None,
    )

    workbook_path = tmp_path / "import.xlsx"
    _write_excel(workbook_path, [("A1", "10"), ("B2", "20")])

    runner = CliRunner()
    result = runner.invoke(
        history_group,
        ["import-table"],
        obj={"config_dir": tmp_path},
        input=f"1\n{workbook_path}\ny\nData\n",
    )

    assert result.exit_code == 0
    assert "Imported 2 row(s)" in result.output
    columns, rows = QueryHistoryDatabase(tmp_path)._load_query_rows(saved.sql)
    assert columns == ["Office", "Value"]
    assert rows == [("A1", "10"), ("B2", "20")]


def test_import_table_command_append(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("SPEC_KIT_ALLOW_UNSAFE_EXPORTS", "1")
    library = SavedQueryLibrary(tmp_path)
    saved = library.save_query(name="monthly", sql="SELECT Office, Value FROM sample")

    db = QueryHistoryDatabase(tmp_path)
    db.record_query_result(
        saved.sql,
        ["Office", "Value"],
        [{"Office": "A1", "Value": "10"}],
        source="test",
        metadata=None,
    )

    workbook_path = tmp_path / "import.xlsx"
    _write_excel(workbook_path, [("B2", "20")], sheet_name="Sheet2")

    runner = CliRunner()
    result = runner.invoke(
        history_group,
        ["import-table"],
        obj={"config_dir": tmp_path},
        input=f"1\n{workbook_path}\nn\nSheet2\n",
    )

    assert result.exit_code == 0
    assert "Imported 1 row(s)" in result.output
    columns, rows = QueryHistoryDatabase(tmp_path)._load_query_rows(saved.sql)
    assert rows == [("A1", "10"), ("B2", "20")]



def test_list_tables_command(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("SPEC_KIT_ALLOW_UNSAFE_EXPORTS", "1")
    library = SavedQueryLibrary(tmp_path)
    saved = library.save_query(name="monthly", sql="SELECT Office, Value FROM sample")
    db = QueryHistoryDatabase(tmp_path)
    db.record_query_result(saved.sql, ["Office", "Value"], [{"Office": "A1", "Value": "10"}], source="test", metadata=None)
    runner = CliRunner()
    result = runner.invoke(history_group, ["list-tables"], obj={"config_dir": tmp_path})
    assert result.exit_code == 0
    assert "History tables" in result.output
    assert "monthly" in result.output

def test_list_tables_show_sql(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("SPEC_KIT_ALLOW_UNSAFE_EXPORTS", "1")
    library = SavedQueryLibrary(tmp_path)
    saved = library.save_query(name="monthly", sql="SELECT " + " OR ".join([f"Col{i}={i}" for i in range(200)]))
    db = QueryHistoryDatabase(tmp_path)
    db.record_query_result(saved.sql, ["Office", "Value"], [{"Office": "A1", "Value": "10"}], source="test", metadata=None)
    runner = CliRunner()
    result = runner.invoke(history_group, ["list-tables", "--show-sql"], obj={"config_dir": tmp_path})
    assert "SQL:" in result.output
    assert "..." in result.output
    assert "truncated" in result.output

def test_table_list_shortcut_invokes_cli(monkeypatch) -> None:
    monkeypatch.setattr(sys, "argv", ["TableList"])
    with patch("opendental_query.cli.shortcuts.cli") as mock_cli:
        table_list_shortcut()
    mock_cli.assert_called_once_with(obj={})


def test_select_history_entry_uses_sanitized_table_for_unlinked_entries(
    tmp_path: Path,
) -> None:
    db = QueryHistoryDatabase(tmp_path)
    db.record_query_result(
        "SELECT 42",
        ["Value"],
        [{"Value": "x"}],
        source="test",
        metadata=None,
    )
    sanitized_table = db.list_queries()[0]["sanitized_table"]
    runner = CliRunner()
    result = runner.invoke(
        history_group,
        ["delete", "--force"],
        obj={"config_dir": tmp_path},
        input="1\n",
    )
    assert result.exit_code == 0
    assert f"  1. {sanitized_table} (unlinked)" in result.output


def test_history_listing_uses_saved_query_alias_when_library_missing(tmp_path: Path) -> None:
    db = QueryHistoryDatabase(tmp_path)
    db.record_query_result(
        "SELECT 1",
        ["Value"],
        [{"Value": "1"}],
        source="test",
        metadata={"saved_query": "monthly"},
    )
    runner = CliRunner()
    result = runner.invoke(history_group, ["list-tables"], obj={"config_dir": tmp_path})
    assert result.exit_code == 0
    assert "1. monthly" in result.output


def test_export_history_prompt_and_alias_selection(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setenv("SPEC_KIT_ALLOW_UNSAFE_EXPORTS", "1")
    db = QueryHistoryDatabase(tmp_path)
    db.record_query_result(
        "SELECT 1",
        ["Value"],
        [{"Value": "1"}],
        source="test",
        metadata={"saved_query": "monthly"},
    )
    runner = CliRunner()
    result = runner.invoke(
        history_group,
        ["export", "--output", str(tmp_path / "exports")],
        obj={"config_dir": tmp_path},
        input="monthly\n",
    )
    assert result.exit_code == 0
    assert "Enter the table name to export" in result.output
    assert "Exported 1 row(s) for 'monthly'" in result.output

def test_export_history_sorts_by_run_then_office(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setenv("SPEC_KIT_ALLOW_UNSAFE_EXPORTS", "1")
    library = SavedQueryLibrary(tmp_path)
    saved = library.save_query(name="monthly", sql="SELECT Office, Value FROM sample")

    db = QueryHistoryDatabase(tmp_path)
    db.record_query_result(
        saved.sql,
        ["Office", "Value"],
        [{"Office": "C1", "Value": "30"}, {"Office": "A1", "Value": "10"}],
        source="test",
        metadata=None,
    )
    db.record_query_result(
        saved.sql,
        ["Office", "Value"],
        [{"Office": "B2", "Value": "40"}, {"Office": "A2", "Value": "20"}],
        source="test",
        metadata=None,
    )

    runner = CliRunner()
    result = runner.invoke(
        history_group,
        ["export", "--saved-query", "monthly", "--output", str(tmp_path / "exports")],
        obj={"config_dir": tmp_path},
    )
    assert result.exit_code == 0
    match = re.search(r"Exported \d+ row\(s\) .* to (.+)\.", result.output)
    assert match
    exported_path = Path(match.group(1).strip())
    workbook = load_workbook(exported_path)
    sheet = workbook.active
    values = list(sheet.iter_rows(values_only=True))
    workbook.close()

    # header + 4 rows expected
    assert [values[1][0], values[2][0], values[3][0], values[4][0]] == ["A1", "C1", "A2", "B2"]


def test_delete_history_removes_table(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("SPEC_KIT_ALLOW_UNSAFE_EXPORTS", "1")
    library = SavedQueryLibrary(tmp_path)
    saved = library.save_query(name="monthly", sql="SELECT Office, Value FROM sample")

    db = QueryHistoryDatabase(tmp_path)
    db.record_query_result(
        saved.sql,
        ["Office", "Value"],
        [{"Office": "A1", "Value": "10"}],
        source="test",
        metadata=None,
    )

    runner = CliRunner()
    result = runner.invoke(
        history_group,
        ["delete", "--saved-query", "monthly", "--force"],
        obj={"config_dir": tmp_path},
    )

    assert result.exit_code == 0
    assert "Deleted history table" in result.output
    assert QueryHistoryDatabase(tmp_path).list_queries() == []


def test_delete_history_interactive(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("SPEC_KIT_ALLOW_UNSAFE_EXPORTS", "1")
    library = SavedQueryLibrary(tmp_path)
    saved = library.save_query(name="monthly", sql="SELECT Office, Value FROM sample")
    db = QueryHistoryDatabase(tmp_path)
    db.record_query_result(
        saved.sql,
        ["Office", "Value"],
        [{"Office": "A1", "Value": "10"}],
        source="test",
        metadata=None,
    )

    runner = CliRunner()
    result = runner.invoke(
        history_group,
        ["delete"],
        obj={"config_dir": tmp_path},
        input="1\ny\n",
    )

    assert result.exit_code == 0
    assert "Deleted history table" in result.output
    assert QueryHistoryDatabase(tmp_path).list_queries() == []


def test_table_delete_shortcut_invokes_cli(monkeypatch) -> None:
    monkeypatch.setattr(sys, "argv", ["TableDelete"])
    with patch("opendental_query.cli.shortcuts.cli") as mock_cli:
        table_delete_shortcut()

    mock_cli.assert_called_once_with(obj={})
    assert sys.argv[1:3] == ["history", "delete"]



def test_table_import_shortcut_invokes_cli(monkeypatch) -> None:
    monkeypatch.setattr(sys, "argv", ["TableImport"])
    with patch("opendental_query.cli.shortcuts.cli") as mock_cli:
        table_import_shortcut()

    mock_cli.assert_called_once_with(obj={})
    assert sys.argv[1:3] == ["history", "import-table"]
